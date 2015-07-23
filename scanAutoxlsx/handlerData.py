#!/usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'mugbya'

import os
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook

def wirteContent(file, data, cell):
    wb = load_workbook(file)
    ew = ExcelWriter(wb)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    ws.cell(cell).value = data
    ws.title = os.path.splitext(file)[0]
    ew.save(file)
